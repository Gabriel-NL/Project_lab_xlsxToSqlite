import sqlite3
import openpyxl
 # Connect to SQLite database
database = sqlite3.connect('data.sqlite')
excel_file = "Producao por municipio (Formatada).xlsx"
excel_file2 = "Produção extrativa vegetal (Formatada).xlsx"



#Functions
def get_first_row_values(sheet):
    try:
        # Get values of the first row as strings
        first_row_values = [str(cell.value).replace(' ', '_') for cell in sheet[1]]
        return first_row_values
    except Exception as e:
        print("An error occurred:", e)
        return None  # Return None if an error occurs


def create_or_update_table(conn, sheet, columns):
    try:
        cursor = conn.cursor()

        # Get the name of the sheet
        sheet_name = sheet.title
        print(f"Name of sheet: {sheet_name}")
        # Check if the table already exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (sheet_name,))
        table_exists = cursor.fetchone()

        if not table_exists:
            print("Table does not exist")
            # If table does not exist, create it with provided columns
            create_table_query = f"CREATE TABLE {sheet_name} ({', '.join(columns)});"
            cursor.execute(create_table_query)
            print("Created Table")
            print(f"Table '{sheet_name}' created with columns: {', '.join(columns)}")

        else:
            print("Table exists")
            # If table exists, fetch its existing columns
            cursor.execute(f"PRAGMA table_info({sheet_name})")
            existing_columns = [row[1] for row in cursor.fetchall()]

            # Check for missing columns and add them
            missing_columns = [column for column in columns if column not in existing_columns]
            if missing_columns:
                for column in missing_columns:
                    print(f"Adding column: {column}")
                    cursor.execute("ALTER TABLE {} ADD COLUMN {} TEXT".format(sheet_name, column))
                    print("Column '{}' added to table '{}'".format(column, sheet_name))
            else:
                print("Table '{}' already exists with matching columns".format(sheet_name))

        # Commit changes to the database
        conn.commit()
    except Exception as e:
        print("An error occurred:", e)



def import_excel_to_database(excel_file, sheet_name, database):
    # Connect to SQLite database
    cursor = database.cursor()

    # Load Excel workbook and select the specified sheet
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb[sheet_name]

    iteration =1
    num_rows = sheet.max_row-1
    # Iterate over rows in the Excel sheet (starting from the second row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extract values from each column
        values = {}
        for col, value in zip(sheet[1], row):
            values[col.value] = value

        # Prepare SQL query to insert values into the database
        columns = ', '.join(values.keys())
        placeholders = ', '.join('?' * len(values))
        query = f"INSERT INTO {sheet_name} ({columns}) VALUES ({placeholders})"

        progress_percentage = (iteration) / num_rows * 100
        print(f"\rProcessing line {iteration}/{num_rows} - Progress: {progress_percentage:.2f}%", end="", flush=True)
        iteration+=1
        # Execute the query and commit changes
        cursor.execute(query, tuple(values.values()))
        database.commit()

    # Close database connection
    #database.close()



sheet_name = "Produção_pecuaria"
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]
columns=get_first_row_values(sheet)
create_or_update_table(database, sheet,columns)
import_excel_to_database(excel_file, sheet_name, database)

sheet_name = "Produção_silvicola"
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]
columns=get_first_row_values(sheet)
create_or_update_table(database, sheet,columns)
import_excel_to_database(excel_file, sheet_name, database)

sheet_name = "Producao_agricola"
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]
columns=get_first_row_values(sheet)
create_or_update_table(database, sheet,columns)
import_excel_to_database(excel_file, sheet_name, database)

sheet_name = "Produção_aquícola"
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]
columns=get_first_row_values(sheet)
create_or_update_table(database, sheet,columns)
import_excel_to_database(excel_file, sheet_name, database)

sheet_name = "Produção_extrativa_vegetal"
wb = openpyxl.load_workbook(excel_file2)
sheet = wb[sheet_name]
columns=get_first_row_values(sheet)
create_or_update_table(database, sheet,columns)
import_excel_to_database(excel_file2, sheet_name, database)

database.close()